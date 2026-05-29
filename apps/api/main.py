from fastapi import Depends, FastAPI, HTTPException, Query, BackgroundTasks, Response, Security
from fastapi.responses import HTMLResponse, FileResponse
from fastapi.security import OAuth2PasswordBearer
from contextlib import asynccontextmanager
from sqlalchemy.orm import Session
from sqlalchemy import func
import os
import threading
import jwt
from redis import Redis
from rq import Queue
from datetime import datetime, timedelta, timezone

from belzakupki_db.read import (
    get_tender,
    list_matches,
    list_tenders,
    serialize_match,
    serialize_tender,
)
from belzakupki_db.session import get_session
from belzakupki_db.models import SearchProfile, NotificationChannel, Tender, TenderMatch, NotificationLog, Tenant, User, CrmConfig, TenderDocument, TenderChatHistory
from belzakupki_db.enums import MatchStatus, NotificationStatus
from belzakupki_db.presets import PRESETS
from belzakupki_db.auth_utils import hash_password, verify_password
from apps.api.schemas import (
    SearchProfileCreate,
    SearchProfileUpdate,
    SearchProfileResponse,
    NotificationChannelCreate,
    NotificationChannelResponse,
    MatchStatusUpdate,
    UserCreate,
    UserLogin,
    UserResponse,
    TokenResponse,
    TenantResponse,
    CrmConfigCreate,
    CrmConfigResponse,
    ChatMessageCreate,
    ChatMessageResponse,
)
import re

def extract_numeric_value(val_str: str | None) -> float | None:
    if not val_str:
        return None
    # Strip spaces and keep digits, dots, commas
    cleaned = "".join(c for c in val_str if c.isdigit() or c in ".,")
    if not cleaned:
        return None
    if "," in cleaned:
        if "." in cleaned:
            cleaned = cleaned.replace(",", "")
        else:
            cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None

@asynccontextmanager
async def lifespan(app: FastAPI):
    """Application lifecycle — API is stateless: no background threads here.

    The scheduler lives in the worker process (worker/scheduler.py).
    """
    yield


app = FastAPI(title="belzakupki", lifespan=lifespan)

from fastapi.staticfiles import StaticFiles
# Mount static files under /assets for compiled Vue app
app.mount("/assets", StaticFiles(directory="apps/api/static/assets", check_dir=False), name="assets")

# --- Аутентификация ---

JWT_SECRET = os.getenv("API_SECRET_KEY", "fallback-secret-for-dev-use-only-1234567890")
JWT_ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24  # 24 часа

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login", auto_error=False)


def create_access_token(data: dict, expires_delta: timedelta | None = None) -> str:
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now(timezone.utc) + expires_delta
    else:
        expire = datetime.now(timezone.utc) + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGORITHM)


def get_current_user(
    token: str | None = Depends(oauth2_scheme),
    session: Session = Depends(get_session)
) -> User:
    """Извлекает текущего пользователя из JWT токена.

    Если авторизация не пройдена или токен отсутствует, в режиме разработки 
    (когда API_SECRET_KEY не задан в окружении) возвращает первого пользователя 
    из базы данных для обратной совместимости.
    """
    secret = os.getenv("API_SECRET_KEY")
    
    if not secret and not token:
        first_user = session.query(User).first()
        if not first_user:
            # Создаем на лету дефолтного пользователя, если база не была засеяна
            tenant = session.query(Tenant).first()
            if not tenant:
                tenant = Tenant(name="ООО Ромашка")
                session.add(tenant)
                session.flush()
            first_user = User(
                tenant_id=tenant.id,
                email="admin@belzakupki.by",
                hashed_password=hash_password("adminpass"),
                full_name="Администратор",
                role="admin"
            )
            session.add(first_user)
            session.commit()
            session.refresh(first_user)
        return first_user
        
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
        
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        email: str = payload.get("sub")
        if email is None:
            raise HTTPException(status_code=401, detail="Invalid token payload")
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Could not validate credentials")
        
    user = session.query(User).filter(User.email == email).one_or_none()
    if user is None:
        raise HTTPException(status_code=401, detail="User not found")
    if not user.is_active:
        raise HTTPException(status_code=401, detail="User is inactive")
    return user


def get_current_tenant(
    current_user: User = Depends(get_current_user)
) -> Tenant:
    """Извлекает организацию (tenant) текущего пользователя."""
    return current_user.tenant


def get_optional_current_user(
    token: str | None = Depends(oauth2_scheme),
    session: Session = Depends(get_session)
) -> User | None:
    """Опционально извлекает пользователя. Возвращает None, если авторизация отсутствует."""
    if not token:
        return None
    try:
        return get_current_user(token, session)
    except HTTPException:
        return None


def get_optional_current_tenant(
    current_user: User | None = Depends(get_optional_current_user)
) -> Tenant | None:
    """Опционально возвращает организацию (tenant) пользователя."""
    return current_user.tenant if current_user else None


def get_current_admin(
    current_user: User = Depends(get_current_user)
) -> User:
    """Гарантирует, что текущий пользователь является администратором."""
    if current_user.role != "admin":
        raise HTTPException(
            status_code=403,
            detail="Доступ запрещен. Требуются права администратора."
        )
    return current_user



# --- Redis и RQ Очереди (Background Jobs) ---

def get_redis_client() -> Redis:
    redis_url = os.getenv("REDIS_URL", "redis://localhost:6379/0")
    return Redis.from_url(redis_url)

def get_task_status_from_redis(tenant_id: int | None = None) -> dict[str, str]:
    try:
        r = get_redis_client()
        suffix = tenant_id if tenant_id is not None else 'global'
        ingest_val = r.get(f"belzakupki:task:ingest:{suffix}")
        notify_val = r.get(f"belzakupki:task:notify:{suffix}")
        return {
            "ingest": ingest_val.decode('utf-8') if ingest_val else "idle",
            "notify": notify_val.decode('utf-8') if notify_val else "idle",
        }
    except Exception as e:
        print(f"Error fetching task status from Redis: {e}")
        return {"ingest": "idle", "notify": "idle"}


def require_api_key(key: str | None = Security(oauth2_scheme), session: Session = Depends(get_session)) -> None:
    """Обратная совместимость: проверяет наличие валидного JWT токена."""
    get_current_user(token=key, session=session)


# Состояние фоновых задач
task_status = {
    "ingest": "idle",
    "notify": "idle",
}
status_lock = threading.Lock()


def run_ingest_task(tenant_id: int | None = None):
    """Фоновая задача сбора тендеров (Dev-режим).

    Считывает активные профили текущей организации и запускает сбор.
    """
    global task_status
    with status_lock:
        if task_status["ingest"] == "running":
            return
        task_status["ingest"] = "running"
    try:
        from belzakupki_db.session import SessionLocal
        from worker.ingest import ingest_goszakupki_tenders, ingest_icetrade_tenders
        with SessionLocal() as session:
            # Считываем активные профили
            query = session.query(SearchProfile).filter(SearchProfile.is_active == True)
            if tenant_id is not None:
                query = query.filter(SearchProfile.tenant_id == tenant_id)
            profiles = query.all()
            # Запускаем динамический сбор
            ingest_goszakupki_tenders(session, profiles=profiles, limit=20)
            ingest_icetrade_tenders(session, profiles=profiles, limit=20)
    except Exception as e:
        print(f"Error during background ingest: {e}")
    finally:
        with status_lock:
            task_status["ingest"] = "idle"


def run_notify_task(tenant_id: int | None = None):
    """Фоновая задача рассылки уведомлений (Dev-режим).

    Инициирует отправку уведомлений по всем новым совпадениям.
    """
    global task_status
    with status_lock:
        if task_status["notify"] == "running":
            return
        task_status["notify"] = "running"
    try:
        from belzakupki_db.session import SessionLocal
        from worker.notifications import dispatch_notifications
        with SessionLocal() as session:
            # Dispatch only notifications belonging to profiles of this tenant if specified
            dispatch_notifications(session, tenant_id=tenant_id)
    except Exception as e:
        print(f"Error during background notify: {e}")
    finally:
        with status_lock:
            task_status["notify"] = "idle"


# --- Раздача Фронтенда ---

@app.get("/", response_class=HTMLResponse)
def read_dashboard(response: Response):
    """Раздает главную страницу панели управления (HTML)."""
    file_path = "apps/api/static/index.html"
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Frontend HTML file not found")
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    with open(file_path, "r", encoding="utf-8") as f:
        return f.read()


@app.get("/style.css")
def read_style(response: Response):
    file_path = "apps/api/static/style.css"
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="style.css not found")
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return FileResponse(file_path, media_type="text/css")


@app.get("/app.js")
def read_app_js(response: Response):
    file_path = "apps/api/static/app.js"
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="app.js not found")
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return FileResponse(file_path, media_type="application/javascript")


# --- Регистрация и Авторизация (Auth Endpoints) ---

@app.post("/api/auth/register", response_model=UserResponse)
def register_user(data: UserCreate, session: Session = Depends(get_session)):
    """Регистрирует новую организацию (Tenant) и администратора."""
    existing_user = session.query(User).filter(User.email == data.email).one_or_none()
    if existing_user:
        raise HTTPException(status_code=400, detail="User with this email already exists")

    tenant_name = data.tenant_name or f"Организация {data.email.split('@')[0]}"
    tenant = Tenant(name=tenant_name)
    session.add(tenant)
    session.flush()

    user = User(
        tenant_id=tenant.id,
        email=data.email,
        hashed_password=hash_password(data.password),
        full_name=data.full_name,
        role="manager"
    )
    session.add(user)
    session.commit()
    session.refresh(user)
    return user


@app.post("/api/auth/login", response_model=TokenResponse)
def login_user(data: UserLogin, session: Session = Depends(get_session)):
    """Аутентифицирует пользователя и возвращает JWT токен."""
    user = session.query(User).filter(User.email == data.email).one_or_none()
    if not user or not verify_password(data.password, user.hashed_password):
        raise HTTPException(status_code=400, detail="Invalid email or password")
    if not user.is_active:
        raise HTTPException(status_code=400, detail="User account is deactivated")

    access_token = create_access_token(
        data={"sub": user.email},
        expires_delta=timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    )
    return {"access_token": access_token, "token_type": "bearer"}


@app.get("/api/auth/me", response_model=UserResponse)
def get_me(current_user: User = Depends(get_current_user)):
    """Возвращает информацию о текущем пользователе."""
    return current_user


# --- Статистика ---

@app.get("/api/stats")
def get_stats(
    current_tenant: Tenant = Depends(get_current_tenant),
    session: Session = Depends(get_session)
):
    """Возвращает агрегированную статистику по тендерам, совпадениям и уведомлениям для текущего тенанта."""
    total_tenders = session.query(func.count(Tender.id)).scalar() or 0
    
    total_matches = session.query(func.count(TenderMatch.id)).join(TenderMatch.profile).filter(SearchProfile.tenant_id == current_tenant.id).scalar() or 0
    new_matches = session.query(func.count(TenderMatch.id)).join(TenderMatch.profile).filter(TenderMatch.status == MatchStatus.NEW, SearchProfile.tenant_id == current_tenant.id).scalar() or 0
    processed_matches = session.query(func.count(TenderMatch.id)).join(TenderMatch.profile).filter(TenderMatch.status == MatchStatus.PROCESSED, SearchProfile.tenant_id == current_tenant.id).scalar() or 0
    expired_matches = session.query(func.count(TenderMatch.id)).join(TenderMatch.profile).filter(TenderMatch.status == MatchStatus.EXPIRED, SearchProfile.tenant_id == current_tenant.id).scalar() or 0

    sent_logs = session.query(func.count(NotificationLog.id)).join(NotificationLog.match).join(TenderMatch.profile).filter(NotificationLog.status == NotificationStatus.SENT, SearchProfile.tenant_id == current_tenant.id).scalar() or 0
    error_logs = session.query(func.count(NotificationLog.id)).join(NotificationLog.match).join(TenderMatch.profile).filter(NotificationLog.status == NotificationStatus.ERROR, SearchProfile.tenant_id == current_tenant.id).scalar() or 0

    return {
        "stats": {
            "total_tenders": total_tenders,
            "total_matches": total_matches,
            "new_matches": new_matches,
            "processed_matches": processed_matches,
            "expired_matches": expired_matches,
            "sent_notifications": sent_logs,
            "error_notifications": error_logs,
        },
        "tasks": get_task_status_from_redis(current_tenant.id)
    }


# --- Управление профилями поиска (Search Profiles CRUD) ---

@app.get("/api/profiles", response_model=list[SearchProfileResponse])
def get_profiles(
    limit: int = Query(default=200, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
    current_tenant: Tenant = Depends(get_current_tenant),
    session: Session = Depends(get_session),
):
    """Возвращает список всех поисковых профилей текущего клиента."""
    return (
        session.query(SearchProfile)
        .filter(SearchProfile.tenant_id == current_tenant.id)
        .order_by(SearchProfile.id.asc())
        .limit(limit)
        .offset(offset)
        .all()
    )


@app.get("/api/presets")
def get_presets(current_user: User = Depends(get_current_user)):
    """Возвращает список всех доступных пресетов с их дефолтными настройками."""
    return list(PRESETS.values())


@app.post("/api/profiles", response_model=SearchProfileResponse)
def create_profile(
    data: SearchProfileCreate,
    current_tenant: Tenant = Depends(get_current_tenant),
    session: Session = Depends(get_session),
):
    """Создает новый поисковый профиль."""
    if data.is_active:
        from belzakupki_db.billing import PLAN_LIMITS
        limits = PLAN_LIMITS.get(current_tenant.plan, PLAN_LIMITS["free"])
        active_count = session.query(SearchProfile).filter(
            SearchProfile.tenant_id == current_tenant.id,
            SearchProfile.is_active.is_(True)
        ).count()
        if active_count >= limits["max_active_profiles"]:
            raise HTTPException(
                status_code=403,
                detail=f"Превышен лимит активных профилей для тарифа '{current_tenant.plan}' (максимум: {limits['max_active_profiles']})"
            )

    profile = SearchProfile(
        tenant_id=current_tenant.id,
        name=data.name,
        description=data.description,
        preset_code=data.preset_code,
        niche_description=data.niche_description,
        keywords=data.keywords,
        negative_keywords=data.negative_keywords,
        regions=data.regions,
        categories=data.categories,
        min_score=data.min_score,
        is_active=data.is_active,
        schedule_interval=None if data.schedule_interval == "manual" else data.schedule_interval,
    )
    session.add(profile)
    session.commit()
    session.refresh(profile)
    return profile


@app.put("/api/profiles/{profile_id}", response_model=SearchProfileResponse)
def update_profile(
    profile_id: int,
    data: SearchProfileUpdate,
    current_tenant: Tenant = Depends(get_current_tenant),
    session: Session = Depends(get_session),
):
    """Обновляет параметры существующего поискового профиля."""
    profile = session.query(SearchProfile).filter(
        SearchProfile.id == profile_id,
        SearchProfile.tenant_id == current_tenant.id
    ).one_or_none()
    if not profile:
        raise HTTPException(status_code=404, detail="Profile not found or access denied")
    
    if data.is_active is True and not profile.is_active:
        from belzakupki_db.billing import PLAN_LIMITS
        limits = PLAN_LIMITS.get(current_tenant.plan, PLAN_LIMITS["free"])
        active_count = session.query(SearchProfile).filter(
            SearchProfile.tenant_id == current_tenant.id,
            SearchProfile.is_active.is_(True),
            SearchProfile.id != profile.id
        ).count()
        if active_count >= limits["max_active_profiles"]:
            raise HTTPException(
                status_code=403,
                detail=f"Превышен лимит активных профилей для тарифа '{current_tenant.plan}' (максимум: {limits['max_active_profiles']})"
            )

    if data.name is not None:
        profile.name = data.name
    if data.description is not None:
        profile.description = data.description
    if data.preset_code is not None:
        profile.preset_code = data.preset_code
    if data.niche_description is not None:
        profile.niche_description = data.niche_description
    if data.keywords is not None:
        profile.keywords = data.keywords
    if data.negative_keywords is not None:
        profile.negative_keywords = data.negative_keywords
    if data.regions is not None:
        profile.regions = data.regions
    if data.categories is not None:
        profile.categories = data.categories
    if data.min_score is not None:
        profile.min_score = data.min_score
    if data.is_active is not None:
        profile.is_active = data.is_active
    if data.schedule_interval is not None:
        profile.schedule_interval = None if data.schedule_interval == "manual" else data.schedule_interval
        
    session.commit()
    session.refresh(profile)
    return profile


@app.delete("/api/profiles/{profile_id}")
def delete_profile(
    profile_id: int,
    current_tenant: Tenant = Depends(get_current_tenant),
    session: Session = Depends(get_session),
):
    """Удаляет поисковый профиль по его ID."""
    profile = session.query(SearchProfile).filter(
        SearchProfile.id == profile_id,
        SearchProfile.tenant_id == current_tenant.id
    ).one_or_none()
    if not profile:
        raise HTTPException(status_code=404, detail="Profile not found or access denied")
    
    session.delete(profile)
    session.commit()
    return {"status": "deleted"}


@app.get("/api/billing/status")
def get_billing_status(
    current_tenant: Tenant = Depends(get_current_tenant),
    session: Session = Depends(get_session),
):
    """Возвращает информацию о текущем тарифе, сроке действия подписки, кредитах и лимитах."""
    from belzakupki_db.billing import PLAN_LIMITS, check_and_reset_billing_cycle
    
    check_and_reset_billing_cycle(session, current_tenant)
    session.commit()
    
    limits = PLAN_LIMITS.get(current_tenant.plan, PLAN_LIMITS["free"])
    
    active_profiles_count = session.query(SearchProfile).filter(
        SearchProfile.tenant_id == current_tenant.id,
        SearchProfile.is_active.is_(True)
    ).count()
    
    return {
        "plan": current_tenant.plan,
        "subscription_expires_at": current_tenant.subscription_expires_at.isoformat() if current_tenant.subscription_expires_at else None,
        "ai_credits_used": current_tenant.ai_credits_used,
        "ai_credits_limit": limits["max_ai_credits"],
        "max_active_profiles": limits["max_active_profiles"],
        "active_profiles_count": active_profiles_count,
        "max_channels_per_profile": limits["max_channels_per_profile"],
        "billing_cycle_started_at": current_tenant.billing_cycle_started_at.isoformat(),
    }


@app.get("/api/admin/stats")
def get_admin_stats(
    current_admin: User = Depends(get_current_admin),
    session: Session = Depends(get_session),
):
    """Возвращает системную статистику для администратора."""
    from belzakupki_db.models import Tenant, User, Tender, TenderMatch, TenderSource

    user_count = session.query(User).count()
    tenant_count = session.query(Tenant).count()
    tender_count = session.query(Tender).count()
    match_count = session.query(TenderMatch).count()

    sources = session.query(TenderSource).all()
    source_stats = []
    for s in sources:
        latest_tender = session.query(Tender).filter(Tender.source_id == s.id).order_by(Tender.created_at.desc()).first()
        count = session.query(Tender).filter(Tender.source_id == s.id).count()
        source_stats.append({
            "code": s.code,
            "name": s.name,
            "total_tenders": count,
            "latest_fetch": latest_tender.created_at.isoformat() if latest_tender else None
        })

    tenants_list = []
    tenants = session.query(Tenant).all()
    for t in tenants:
        active_profiles = session.query(SearchProfile).filter(SearchProfile.tenant_id == t.id, SearchProfile.is_active == True).count()
        tenants_list.append({
            "id": t.id,
            "name": t.name,
            "is_active": t.is_active,
            "plan": t.plan,
            "ai_credits_used": t.ai_credits_used,
            "active_profiles": active_profiles,
            "created_at": t.created_at.isoformat() if t.created_at else None
        })

    # Return a log stream of parser checks (as in a real dashboard)
    logs = [
        {"timestamp": datetime.now(timezone.utc).isoformat(), "level": "INFO", "message": "Воркер запущен, ожидание очереди задач..."},
    ]
    for s in sources:
        latest_t = session.query(Tender).filter(Tender.source_id == s.id).order_by(Tender.created_at.desc()).first()
        if latest_t:
            logs.append({
                "timestamp": latest_t.created_at.isoformat(),
                "level": "INFO",
                "message": f"Источник {s.name} успешно опрошен. Обнаружено {latest_tender.id if latest_tender else 0} публикаций."
            })

    return {
        "user_count": user_count,
        "tenant_count": tenant_count,
        "tender_count": tender_count,
        "match_count": match_count,
        "sources": source_stats,
        "tenants": tenants_list,
        "logs": logs
    }


# --- Управление каналами уведомлений (Notification Channels) ---

@app.get("/api/profiles/{profile_id}/channels", response_model=list[NotificationChannelResponse])
def get_profile_channels(
    profile_id: int,
    current_tenant: Tenant = Depends(get_current_tenant),
    session: Session = Depends(get_session),
):
    """Возвращает все настроенные каналы уведомлений для указанного поискового профиля."""
    profile = session.query(SearchProfile).filter(
        SearchProfile.id == profile_id,
        SearchProfile.tenant_id == current_tenant.id
    ).one_or_none()
    if not profile:
        raise HTTPException(status_code=404, detail="Profile not found or access denied")

    return session.query(NotificationChannel).filter(NotificationChannel.profile_id == profile_id).all()


@app.post("/api/profiles/{profile_id}/channels", response_model=NotificationChannelResponse)
def create_or_update_channel(
    profile_id: int,
    data: NotificationChannelCreate,
    current_tenant: Tenant = Depends(get_current_tenant),
    session: Session = Depends(get_session),
):
    """Создает новый или обновляет существующий канал уведомлений для указанного профиля."""
    # Проверяем, существует ли профиль и принадлежит ли он текущему клиенту
    profile = session.query(SearchProfile).filter(
        SearchProfile.id == profile_id,
        SearchProfile.tenant_id == current_tenant.id
    ).one_or_none()
    if not profile:
        raise HTTPException(status_code=404, detail="Profile not found or access denied")
        
    channel = session.query(NotificationChannel).filter(
        NotificationChannel.profile_id == profile_id,
        NotificationChannel.type == data.type
    ).first()
    
    if data.is_active:
        from belzakupki_db.billing import PLAN_LIMITS
        limits = PLAN_LIMITS.get(current_tenant.plan, PLAN_LIMITS["free"])
        
        active_channels_query = session.query(NotificationChannel).filter(
            NotificationChannel.profile_id == profile_id,
            NotificationChannel.is_active.is_(True)
        )
        if channel:
            active_channels_query = active_channels_query.filter(NotificationChannel.id != channel.id)
            
        active_channels_count = active_channels_query.count()
        if active_channels_count >= limits["max_channels_per_profile"]:
            raise HTTPException(
                status_code=403,
                detail=f"Превышен лимит активных каналов уведомлений на профиль для тарифа '{current_tenant.plan}' (максимум: {limits['max_channels_per_profile']})"
            )
    
    if channel:
        channel.name = data.name
        channel.config = data.config
        channel.is_active = data.is_active
    else:
        channel = NotificationChannel(
            profile_id=profile_id,
            type=data.type,
            name=data.name,
            config=data.config,
            is_active=data.is_active,
        )
        session.add(channel)
        
    session.commit()
    session.refresh(channel)
    return channel


# --- Запуск действий (Actions) ---

@app.post("/api/actions/ingest")
def trigger_ingest(
    current_tenant: Tenant = Depends(get_current_tenant),
):
    """Инициирует запуск фонового задания сбора тендеров через очередь RQ."""
    if current_tenant.plan == "free":
        raise HTTPException(status_code=403, detail="Запуск сборов доступен только на платных тарифах")
    r = get_redis_client()
    status_key = f"belzakupki:task:ingest:{current_tenant.id}"
    
    current_status = r.get(status_key)
    if current_status and current_status.decode('utf-8') == "running":
        return {"status": "already_running"}
        
    q = Queue("default", connection=r)
    # Set status immediately in redis to prevent double clicks
    r.set(status_key, "running")
    q.enqueue("worker.tasks.run_ingest_task_job", current_tenant.id)
    return {"status": "started"}


@app.post("/api/actions/notify")
def trigger_notify(
    current_tenant: Tenant = Depends(get_current_tenant),
):
    """Инициирует запуск фонового задания отправки уведомлений через очередь RQ."""
    if current_tenant.plan == "free":
        raise HTTPException(status_code=403, detail="Запуск рассылок доступен только на платных тарифах")
    r = get_redis_client()
    status_key = f"belzakupki:task:notify:{current_tenant.id}"
    
    current_status = r.get(status_key)
    if current_status and current_status.decode('utf-8') == "running":
        return {"status": "already_running"}
        
    q = Queue("default", connection=r)
    # Set status immediately in redis to prevent double clicks
    r.set(status_key, "running")
    q.enqueue("worker.tasks.run_notify_task_job", current_tenant.id)
    return {"status": "started"}


# --- Существующие эндпоинты ---

@app.get("/api/tenders")
def tenders(
    limit: int = Query(default=20, ge=1, le=100),
    offset: int = Query(default=0, ge=0),
    matched_only: bool = False,
    q: str | None = Query(default=None, min_length=1),
    current_tenant: Tenant | None = Depends(get_optional_current_tenant),
    session: Session = Depends(get_session),
):
    """Возвращает список сохраненных тендеров с поддержкой фильтрации и текстового поиска."""
    created_since = None
    if current_tenant is None:
        # Guest anonymous mode: only show today's tenders (last 24 hours)
        from datetime import datetime, timezone, timedelta
        created_since = datetime.now(timezone.utc) - timedelta(hours=24)
        matched_only = False

    items = list_tenders(
        session,
        limit=limit,
        offset=offset,
        matched_only=matched_only,
        query=q,
        created_since=created_since,
    )

    serialized = []
    for item in items:
        data = serialize_tender(item, tenant_id=current_tenant.id if current_tenant else None)
        if current_tenant is None:
            # Strip files, lots, contacts, terms and ИИ for guest preview
            data["attachments"] = []
            data["lots"] = []
            data["contacts"] = None
            data["delivery_terms"] = None
            data["payment_terms"] = None
            data["ai_relevance"] = None
            data["ai_analysis"] = None
        serialized.append(data)

    return {
        "items": serialized,
        "limit": limit,
        "offset": offset,
    }


@app.get("/api/tenders/{tender_id}")
def tender(
    tender_id: int,
    current_tenant: Tenant = Depends(get_current_tenant),
    session: Session = Depends(get_session),
):
    """Возвращает детальную информацию о конкретном тендере по его ID."""
    item = get_tender(session, tender_id)

    if item is None:
        raise HTTPException(status_code=404, detail="Tender not found")

    return serialize_tender(item, tenant_id=current_tenant.id)


@app.get("/api/matches")
def matches(
    limit: int = Query(default=20, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
    profile_id: int | None = Query(default=None),
    status: str | None = Query(default=None),
    current_tenant: Tenant = Depends(get_current_tenant),
    session: Session = Depends(get_session),
):
    """Возвращает список совпадений тендеров с профилями поиска, отсортированный по уровню соответствия."""
    if profile_id is not None:
        prof = session.query(SearchProfile).filter(
            SearchProfile.id == profile_id,
            SearchProfile.tenant_id == current_tenant.id
        ).one_or_none()
        if not prof:
            raise HTTPException(status_code=403, detail="Profile not found or access denied")

    items = list_matches(
        session,
        limit=limit,
        offset=offset,
        profile_id=profile_id,
        status=status,
        tenant_id=current_tenant.id,
    )

    return {
        "items": [serialize_match(item) for item in items],
        "limit": limit,
        "offset": offset,
    }


@app.put("/api/matches/{match_id}/status")
def update_match_status(
    match_id: int,
    data: MatchStatusUpdate,
    current_tenant: Tenant = Depends(get_current_tenant),
    session: Session = Depends(get_session),
):
    """Обновляет статус совпадения тендера (Канбан-доска)."""
    match = session.query(TenderMatch).join(TenderMatch.profile).filter(
        TenderMatch.id == match_id,
        SearchProfile.tenant_id == current_tenant.id
    ).one_or_none()
    if not match:
        raise HTTPException(status_code=404, detail="Match not found or access denied")

    valid_statuses = [status.value for status in MatchStatus]
    if data.status not in valid_statuses:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid status. Must be one of: {', '.join(valid_statuses)}"
        )

    match.status = data.status
    session.commit()
    session.refresh(match)
    return serialize_match(match)


@app.get("/api/analytics/competitors")
def get_competitor_analytics(
    current_tenant: Tenant = Depends(get_current_tenant),
    session: Session = Depends(get_session)
):
    """Возвращает агрегированную аналитику по победителям, заказчикам и снижению цены."""
    from belzakupki_db.models import TenderResult, Tender
    from sqlalchemy import select, func

    # 1. Top Winners
    winner_stmt = (
        select(
            TenderResult.winner_name,
            TenderResult.winner_unp,
            func.count(TenderResult.id).label("wins_count"),
            func.sum(TenderResult.contract_price).label("total_amount")
        )
        .where(TenderResult.winner_name != None)
        .group_by(TenderResult.winner_name, TenderResult.winner_unp)
        .order_by(func.count(TenderResult.id).desc())
        .limit(10)
    )
    winners_res = session.execute(winner_stmt).all()
    top_winners = [
        {
            "name": r.winner_name,
            "unp": r.winner_unp,
            "wins_count": r.wins_count,
            "total_amount": float(r.total_amount) if r.total_amount is not None else 0.0,
        }
        for r in winners_res
    ]

    # 2. Top Customers
    customer_stmt = (
        select(
            Tender.customer_name,
            func.count(Tender.id).label("tenders_count"),
            func.sum(TenderResult.contract_price).label("total_amount")
        )
        .outerjoin(TenderResult, Tender.id == TenderResult.tender_id)
        .where(Tender.customer_name != None)
        .group_by(Tender.customer_name)
        .order_by(func.count(Tender.id).desc())
        .limit(10)
    )
    customers_res = session.execute(customer_stmt).all()
    top_customers = [
        {
            "name": r.customer_name,
            "tenders_count": r.tenders_count,
            "total_amount": float(r.total_amount) if r.total_amount is not None else 0.0,
        }
        for r in customers_res
    ]

    # 3. Average Price Reduction
    stmt = (
        select(TenderResult.contract_price, Tender.raw_data)
        .join(Tender, Tender.id == TenderResult.tender_id)
        .where(TenderResult.contract_price != None)
    )
    results = session.execute(stmt).all()

    total_discount = 0.0
    count = 0
    percentages = []

    for contract_price, raw_data in results:
        if not raw_data:
            continue
        est_str = raw_data.get("estimated_value")
        est_val = extract_numeric_value(est_str)
        con_val = float(contract_price)

        if est_val and est_val > 0 and con_val > 0 and est_val >= con_val:
            discount = est_val - con_val
            pct = (discount / est_val) * 100
            percentages.append(pct)
            total_discount += discount
            count += 1

    avg_discount_pct = sum(percentages) / len(percentages) if percentages else 0.0

    return {
        "top_winners": top_winners,
        "top_customers": top_customers,
        "metrics": {
            "average_discount_percentage": round(avg_discount_pct, 2),
            "total_discount_amount": round(total_discount, 2),
            "analyzed_count": count
        }
    }


@app.get("/api/reports/export/excel")
def export_excel(
    profile_id: int | None = Query(None),
    status: str | None = Query(None),
    current_tenant: Tenant = Depends(get_current_tenant),
    session: Session = Depends(get_session)
):
    """Генерирует Excel-отчет по совпавшим тендерам (с фильтрацией по профилю/статусу)."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    from datetime import datetime

    if profile_id is not None:
        prof = session.query(SearchProfile).filter(
            SearchProfile.id == profile_id,
            SearchProfile.tenant_id == current_tenant.id
        ).one_or_none()
        if not prof:
            raise HTTPException(status_code=403, detail="Profile not found or access denied")

    items = list_matches(session, limit=5000, profile_id=profile_id, status=status, tenant_id=current_tenant.id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Тендеры"

    ws.views.sheetView[0].showGridLines = True

    headers = [
        "ID совпадения",
        "ID тендера",
        "Название тендера",
        "Источник",
        "Ссылка",
        "Заказчик",
        "Ориентировочная стоимость",
        "Дата публикации",
        "Дедлайн",
        "Балл соответствия",
        "Статус проработки",
        "ИИ-Релевантность",
        "ИИ-Анализ (DeepSeek)",
        "Победитель закупки",
        "Сумма договора"
    ]

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E1E24", end_color="1E1E24", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center", wrap_text=True)

    thin_side = Side(border_style="thin", color="CCCCCC")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border_all

    for item in items:
        tender = item.tender
        source_code = tender.source.code if tender.source else ""
        result = tender.result

        raw_data = tender.raw_data or {}
        est_val = raw_data.get("estimated_value", "")
        published_at_str = tender.published_at.strftime("%d.%m.%Y %H:%M") if tender.published_at else ""
        deadline_at_str = tender.deadline_at.strftime("%d.%m.%Y %H:%M") if tender.deadline_at else ""

        ai_rel = "Да" if item.ai_relevance else ("Нет" if item.ai_relevance is False else "Не проводился")
        ai_summary = ""
        if item.ai_analysis:
            ai_summary = item.ai_analysis.get("relevance_explanation") or item.reason or ""

        winner = result.winner_name if result else ""
        price = ""
        if result and result.contract_price is not None:
            price = f"{result.contract_price} {result.currency or 'BYN'}"

        row_data = [
            item.id,
            tender.id,
            tender.title,
            source_code,
            tender.url,
            tender.customer_name or "",
            est_val,
            published_at_str,
            deadline_at_str,
            float(item.score),
            item.status,
            ai_rel,
            ai_summary,
            winner,
            price
        ]

        ws.append(row_data)
        row_idx = ws.max_row

        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border_all
            cell.font = Font(name="Calibri", size=10)

            if col_idx in [1, 2, 4, 8, 9, 10, 11, 12]:
                cell.alignment = center_align
            elif col_idx in [7, 15]:
                cell.alignment = right_align
            else:
                cell.alignment = left_align

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = min(len(val_str), 40)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"tenders_report_{timestamp}.xlsx"

    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


# --- CRM Integrations (Phase 4) ---

@app.get("/api/crm/settings")
def get_crm_settings(
    current_tenant: Tenant = Depends(get_current_tenant),
    session: Session = Depends(get_session),
):
    """Возвращает настройки интеграции с CRM для текущего арендатора."""
    configs = session.query(CrmConfig).filter(CrmConfig.tenant_id == current_tenant.id).all()
    
    # Mask api_token before sending to frontend
    res = []
    for c in configs:
        res.append({
            "id": c.id,
            "tenant_id": c.tenant_id,
            "crm_type": c.crm_type,
            "is_active": c.is_active,
            "webhook_url": c.webhook_url,
            "subdomain": c.subdomain,
            "api_token": "********" if c.api_token else None,
            "custom_mappings": c.custom_mappings,
        })
    return res


@app.post("/api/crm/settings")
def save_crm_settings(
    data: CrmConfigCreate,
    current_tenant: Tenant = Depends(get_current_tenant),
    session: Session = Depends(get_session),
):
    """Создает или обновляет настройки интеграции с CRM."""
    # Если эта интеграция активируется, деактивируем все остальные для этого tenant
    if data.is_active:
        session.query(CrmConfig).filter(
            CrmConfig.tenant_id == current_tenant.id,
            CrmConfig.crm_type != data.crm_type
        ).update({"is_active": False})
        
    config = session.query(CrmConfig).filter(
        CrmConfig.tenant_id == current_tenant.id,
        CrmConfig.crm_type == data.crm_type
    ).first()
    
    if config:
        config.is_active = data.is_active
        config.webhook_url = data.webhook_url
        config.subdomain = data.subdomain
        config.custom_mappings = data.custom_mappings
        # Обновляем токен, только если он передан и не является плейсхолдером
        if data.api_token and data.api_token != "********":
            config.api_token = data.api_token
    else:
        token_val = data.api_token if data.api_token != "********" else None
        config = CrmConfig(
            tenant_id=current_tenant.id,
            crm_type=data.crm_type,
            is_active=data.is_active,
            webhook_url=data.webhook_url,
            subdomain=data.subdomain,
            api_token=token_val,
            custom_mappings=data.custom_mappings,
        )
        session.add(config)
        
    session.commit()
    session.refresh(config)
    
    return {
        "id": config.id,
        "tenant_id": config.tenant_id,
        "crm_type": config.crm_type,
        "is_active": config.is_active,
        "webhook_url": config.webhook_url,
        "subdomain": config.subdomain,
        "api_token": "********" if config.api_token else None,
        "custom_mappings": config.custom_mappings,
    }


@app.post("/api/crm/settings/test")
async def test_crm_settings(
    data: CrmConfigCreate,
    current_tenant: Tenant = Depends(get_current_tenant),
    session: Session = Depends(get_session),
):
    """Тестирует отправку лида/сделки с указанными настройками."""
    from apps.api.services.crm_service import export_to_bitrix24, export_to_amocrm
    
    # Создаем mock-объекты для теста
    class MockTender:
        title = "ТЕСТ: Проверка интеграции BelZakupki"
        customer_name = "Тестовый Заказчик ООО"
        url = "https://goszakupki.by"
        deadline_at = datetime.now(timezone.utc) + timedelta(days=5)
        raw_data = {"estimated_value": "100 000 BYN"}
        
    class MockMatch:
        score = 95
        matched_keywords = ["тест", "интеграция"]
        reason = "Тестовое соответствие"
        ai_analysis = {
            "relevance_explanation": "Тестовый детальный анализ ИИ (DeepSeek).",
            "key_points": ["Пункт спецификации 1", "Пункт спецификации 2"],
            "risks": ["Тестовый риск по отсрочке платежа 90 дней"],
            "commercial_proposal_info": {
                "scope": "Тестовый объем поставки оборудования",
                "requirements": "Тестовые требования к опыту работы",
                "budget_notes": "Ориентировочный бюджет подтвержден"
            }
        }
        
    try:
        if data.crm_type == "bitrix24":
            if not data.webhook_url:
                raise HTTPException(status_code=400, detail="Webhook URL обязателен для Битрикс24")
            deal_id = await export_to_bitrix24(data.webhook_url, MockTender(), MockMatch())
            return {"success": True, "deal_id": deal_id}
            
        elif data.crm_type == "amocrm":
            if not data.subdomain:
                raise HTTPException(status_code=400, detail="Субдомен обязателен для amoCRM")
            token = data.api_token
            if token == "********":
                db_conf = session.query(CrmConfig).filter(
                    CrmConfig.tenant_id == current_tenant.id,
                    CrmConfig.crm_type == "amocrm"
                ).first()
                if not db_conf or not db_conf.api_token:
                    raise HTTPException(status_code=400, detail="Токен авторизации не найден в сохраненной конфигурации")
                token = db_conf.api_token
            if not token:
                raise HTTPException(status_code=400, detail="Токен авторизации обязателен для amoCRM")
                
            deal_id = await export_to_amocrm(data.subdomain, token, MockTender(), MockMatch())
            return {"success": True, "deal_id": deal_id}
            
        else:
            raise HTTPException(status_code=400, detail="Неподдерживаемый тип CRM")
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/tenders/matches/{match_id}/export-crm")
async def export_match_to_crm(
    match_id: int,
    current_tenant: Tenant = Depends(get_current_tenant),
    session: Session = Depends(get_session),
):
    """Экспортирует найденное совпадение в активную CRM систему и переводит статус в 'В работе'."""
    from apps.api.services.crm_service import export_to_bitrix24, export_to_amocrm
    
    match = session.query(TenderMatch).join(TenderMatch.profile).filter(
        TenderMatch.id == match_id,
        SearchProfile.tenant_id == current_tenant.id
    ).one_or_none()
    
    if not match:
        raise HTTPException(status_code=404, detail="Совпадение не найдено или доступ запрещен")
        
    crm_conf = session.query(CrmConfig).filter(
        CrmConfig.tenant_id == current_tenant.id,
        CrmConfig.is_active == True
    ).first()
    
    if not crm_conf:
        raise HTTPException(
            status_code=400,
            detail="Активная интеграция с CRM не настроена. Пожалуйста, включите и настройте интеграцию."
        )
        
    try:
        if crm_conf.crm_type == "bitrix24":
            if not crm_conf.webhook_url:
                raise Exception("В настройках Битрикс24 отсутствует Webhook URL")
            deal_id = await export_to_bitrix24(crm_conf.webhook_url, match.tender, match)
            
        elif crm_conf.crm_type == "amocrm":
            if not crm_conf.subdomain or not crm_conf.api_token:
                raise Exception("В настройках amoCRM отсутствуют субдомен или токен доступа")
            deal_id = await export_to_amocrm(crm_conf.subdomain, crm_conf.api_token, match.tender, match)
            
        else:
            raise Exception(f"Неизвестный тип CRM: {crm_conf.crm_type}")
            
        match.crm_deal_id = deal_id
        match.status = MatchStatus.IN_WORK.value
        session.commit()
        session.refresh(match)
        return serialize_match(match)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/matches/{match_id}/chat", response_model=list[ChatMessageResponse])
def get_match_chat_history(
    match_id: int,
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    """Возвращает историю чата с ИИ по конкретному совпадению."""
    if current_user.tenant.plan == "free":
        raise HTTPException(status_code=403, detail="ИИ чат-ассистент доступен только на платных тарифах")
    match = session.query(TenderMatch).join(TenderMatch.profile).filter(
        TenderMatch.id == match_id,
        SearchProfile.tenant_id == current_user.tenant_id
    ).one_or_none()
    
    if not match:
        raise HTTPException(status_code=404, detail="Совпадение не найдено или доступ запрещен")
        
    history = session.query(TenderChatHistory).filter(
        TenderChatHistory.match_id == match_id
    ).order_by(TenderChatHistory.created_at.asc()).all()
    
    return history


def retrieve_relevant_context(full_text: str, query: str, max_chars: int = 80000) -> str:
    """Извлекает наиболее релевантные абзацы из спецификаций, если объем текста слишком большой."""
    if len(full_text) <= max_chars:
        return full_text
        
    # Разбиваем текст на параграфы
    paragraphs = [p.strip() for p in re.split(r'\n\n+|\n(?=\s*-|\s*\d+\.)', full_text) if p.strip()]
    
    query_words = [w.lower() for w in re.findall(r'\w+', query) if len(w) > 2]
    if not query_words:
        # Возвращаем первые абзацы до лимита
        res = []
        total_len = 0
        for p in paragraphs:
            if total_len + len(p) < max_chars:
                res.append(p)
                total_len += len(p)
            else:
                break
        return "\n\n".join(res)
        
    scored_paragraphs = []
    for p in paragraphs:
        # Подсчет вхождений ключевых слов вопроса
        score = sum(p.lower().count(qw) for qw in query_words)
        scored_paragraphs.append((score, p))
        
    # Сортируем по релевантности
    scored_paragraphs.sort(key=lambda x: x[0], reverse=True)
    
    selected = []
    total_len = 0
    for score, p in scored_paragraphs:
        if total_len + len(p) < max_chars:
            selected.append(p)
            total_len += len(p)
        else:
            break
            
    return "\n\n".join(selected)


@app.post("/api/matches/{match_id}/chat", response_model=ChatMessageResponse)
async def post_match_chat_message(
    match_id: int,
    data: ChatMessageCreate,
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    """Отправляет сообщение ИИ-ассистенту и возвращает его ответ в контексте документов тендера."""
    if current_user.tenant.plan == "free":
        raise HTTPException(status_code=403, detail="ИИ чат-ассистент доступен только на платных тарифах")
    import httpx
    import logging
    logger = logging.getLogger("belzakupki.chat")
    
    match = session.query(TenderMatch).join(TenderMatch.profile).filter(
        TenderMatch.id == match_id,
        SearchProfile.tenant_id == current_user.tenant_id
    ).one_or_none()
    
    if not match:
        raise HTTPException(status_code=404, detail="Совпадение не найдено или доступ запрещен")
        
    # 1. Сохраняем сообщение пользователя
    user_msg = TenderChatHistory(
        match_id=match_id,
        user_id=current_user.id,
        role="user",
        message=data.message
    )
    session.add(user_msg)
    session.commit()
    
    # 2. Собираем контекст из документов тендера
    docs = session.query(TenderDocument).filter(
        TenderDocument.tender_id == match.tender_id
    ).all()
    
    context_parts = []
    for doc in docs:
        context_parts.append(f"--- Файл спецификации: {doc.file_name} ---\n{doc.content}")
        
    full_context = "\n\n".join(context_parts)
    if not full_context:
        # Фолбэк на название и описание
        full_context = f"Название закупки: {match.tender.title}\nОписание: {match.tender.description or 'Описание отсутствует'}"
        
    # Обрезаем контекст, если он слишком большой
    relevant_context = retrieve_relevant_context(full_context, data.message)
    
    # 3. Достаем историю чата
    chat_history = session.query(TenderChatHistory).filter(
        TenderChatHistory.match_id == match_id
    ).order_by(TenderChatHistory.created_at.asc()).all()
    
    # Формируем промпт для DeepSeek
    messages = [
        {"role": "system", "content": (
            "Вы — опытный ИИ-ассистент тендерного отдела платформы BelZakupki.\n"
            "Ваша задача — анализировать требования, спецификации и документы тендера, предоставленные в контексте, "
            "и отвечать на вопросы менеджера по продажам.\n"
            "Отвечайте на русском языке, лаконично, структурированно (используйте списки и выделения) и строго опирайтесь на факты из документов. "
            "Если ответа нет в документах, прямо напишите об этом.\n\n"
            f"ТЕКСТ ДОКУМЕНТОВ ТЕНДЕРА:\n{relevant_context}"
        )}
    ]
    
    # Добавляем историю (исключая последнее сообщение пользователя, так как мы добавим его ниже)
    for msg in chat_history[:-1]:
        messages.append({"role": msg.role, "content": msg.message})
        
    # Добавляем последнее сообщение пользователя
    messages.append({"role": "user", "content": data.message})
    
    token = os.getenv("DEEPSEEK_TOKEN")
    if not token or token == "your-deepseek-token":
        # Фолбэк для разработки
        answer = f"ИИ-ассистент (режим симуляции): Вы спросили '{data.message}'. К сожалению, DEEPSEEK_TOKEN не настроен."
    else:
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        payload = {
            "model": "deepseek-chat",
            "messages": messages,
            "temperature": 0.2
        }
        
        async with httpx.AsyncClient(timeout=45.0) as client:
            try:
                response = await client.post("https://api.deepseek.com/chat/completions", json=payload, headers=headers)
                response.raise_for_status()
                res_json = response.json()
                answer = res_json["choices"][0]["message"]["content"]
            except Exception as e:
                logger.error(f"DeepSeek Chat completions failed: {e}")
                # Фолбэк на случай сбоя API
                answer = f"Извините, не удалось получить ответ от ИИ-ассистента из-за технической ошибки: {str(e)}"
                
    # 4. Сохраняем ответ ассистента в БД
    assistant_msg = TenderChatHistory(
        match_id=match_id,
        user_id=current_user.id,
        role="assistant",
        message=answer
    )
    session.add(assistant_msg)
    session.commit()
    session.refresh(assistant_msg)
    
    return assistant_msg


@app.get("/{catchall:path}", response_class=HTMLResponse)
def catch_all(catchall: str, response: Response):
    """Возвращает index.html для всех роутов SPA (Vite/Vue), кроме API и статических файлов."""
    if (
        catchall.startswith("api") 
        or catchall.startswith("assets") 
        or catchall.endswith(".js") 
        or catchall.endswith(".css") 
        or catchall.endswith(".svg")
        or catchall.endswith(".png")
        or catchall.endswith(".ico")
    ):
        raise HTTPException(status_code=404, detail="Not Found")
        
    file_path = "apps/api/static/index.html"
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Frontend HTML file not found")
        
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    with open(file_path, "r", encoding="utf-8") as f:
        return f.read()


